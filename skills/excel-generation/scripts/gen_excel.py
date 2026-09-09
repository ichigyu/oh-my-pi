#!/usr/bin/env python3
"""从 JSON 输入生成 .xlsx 文件。

JSON 输入来自 stdin，或通过 --input <path> 传入文件。
输出结构化 JSON 到 stdout：{"ok": true, "output": "..."} 或 {"ok": false, "error": {...}}。
"""

import argparse
import json
import os
import sys
import traceback


def error_out(category, message, detail=None):
    payload = {
        "ok": False,
        "error": {
            "category": category,
            "message": message,
        },
    }
    if detail is not None:
        payload["error"]["detail"] = detail
    print(json.dumps(payload, ensure_ascii=False))
    sys.exit(1)


try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    error_out("dependency_error", f"openpyxl 依赖不可用: {exc}")


def build_font(style):
    kwargs = {}
    if "bold" in style:
        kwargs["bold"] = bool(style["bold"])
    if "italic" in style:
        kwargs["italic"] = bool(style["italic"])
    if "font_size" in style:
        kwargs["size"] = style["font_size"]
    if "font_color" in style:
        kwargs["color"] = style["font_color"]
    if not kwargs:
        return None
    return Font(**kwargs)


def build_fill(style):
    color = style.get("fill_color")
    if not color:
        return None
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def build_alignment(style):
    align = style.get("align")
    valign = style.get("valign")
    if not align and not valign:
        return None
    kwargs = {}
    if align:
        kwargs["horizontal"] = align
    if valign:
        kwargs["vertical"] = valign
    return Alignment(**kwargs)


def build_border(style):
    spec = style.get("border")
    if not spec:
        return None
    style_name = spec if isinstance(spec, str) else "thin"
    side = Side(style=style_name)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_style(cell, style):
    if not style:
        return
    font = build_font(style)
    if font:
        cell.font = font
    fill = build_fill(style)
    if fill:
        cell.fill = fill
    alignment = build_alignment(style)
    if alignment:
        cell.alignment = alignment
    border = build_border(style)
    if border:
        cell.border = border
    if "number_format" in style:
        cell.number_format = style["number_format"]


CHART_TYPES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
}


def parse_ref(ws, ref_spec):
    return Reference(
        ws,
        min_col=ref_spec["min_col"],
        min_row=ref_spec["min_row"],
        max_col=ref_spec.get("max_col", ref_spec["min_col"]),
        max_row=ref_spec.get("max_row", ref_spec["min_row"]),
    )


def add_chart(ws, chart_spec):
    chart_type = chart_spec.get("type")
    chart_cls = CHART_TYPES.get(chart_type)
    if chart_cls is None:
        raise ValueError(f"unsupported chart type: {chart_type} (支持 {list(CHART_TYPES)})")

    chart = chart_cls()
    if chart_spec.get("title"):
        chart.title = chart_spec["title"]

    data_spec = chart_spec.get("data")
    if not data_spec:
        raise ValueError("chart spec missing required field: data")
    data_ref = parse_ref(ws, data_spec)
    chart.add_data(data_ref, titles_from_data=bool(chart_spec.get("titles_from_data", True)))

    categories_spec = chart_spec.get("categories")
    if categories_spec:
        chart.set_categories(parse_ref(ws, categories_spec))

    anchor = chart_spec.get("anchor", "E2")
    ws.add_chart(chart, anchor)


def add_conditional_formatting(ws, rule_spec):
    cell_range = rule_spec.get("range")
    if not cell_range:
        raise ValueError("conditional_formatting spec missing required field: range")

    rule_type = rule_spec.get("type")
    fill = build_fill(rule_spec) or PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    if rule_type == "color_scale":
        colors = rule_spec.get("colors", ["F8696B", "FFEB84", "63BE7B"])
        rule = ColorScaleRule(
            start_type="min", start_color=colors[0],
            mid_type="percentile", mid_value=50, mid_color=colors[1],
            end_type="max", end_color=colors[2],
        )
    elif rule_type == "cell_is":
        rule = CellIsRule(
            operator=rule_spec.get("operator", "greaterThan"),
            formula=rule_spec.get("formula", ["0"]),
            fill=fill,
        )
    elif rule_type == "formula":
        rule = FormulaRule(formula=rule_spec.get("formula", []), fill=fill)
    else:
        raise ValueError(f"unsupported conditional_formatting type: {rule_type}")

    ws.conditional_formatting.add(cell_range, rule)


def add_data_validation(ws, dv_spec):
    cell_range = dv_spec.get("range")
    if not cell_range:
        raise ValueError("data_validation spec missing required field: range")

    dv = DataValidation(
        type=dv_spec.get("type", "list"),
        formula1=dv_spec.get("formula1"),
        allow_blank=bool(dv_spec.get("allow_blank", True)),
    )
    if dv_spec.get("error_message") or dv_spec.get("error_title"):
        dv.error = dv_spec.get("error_message")
        dv.errorTitle = dv_spec.get("error_title", "无效输入")
    ws.add_data_validation(dv)
    dv.add(cell_range)


def write_sheet(ws, sheet_spec):
    rows = sheet_spec.get("rows", [])
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, cell_spec in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell_spec, dict):
                cell.value = cell_spec.get("value")
                apply_style(cell, cell_spec.get("style"))
            else:
                cell.value = cell_spec

    for col_spec in sheet_spec.get("columns", []):
        idx = col_spec.get("index")
        width = col_spec.get("width")
        if idx is None or width is None:
            continue
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row_num_str, height in sheet_spec.get("row_heights", {}).items():
        ws.row_dimensions[int(row_num_str)].height = height

    if sheet_spec.get("freeze_header"):
        ws.freeze_panes = "A2"

    for chart_spec in sheet_spec.get("charts", []):
        add_chart(ws, chart_spec)

    for rule_spec in sheet_spec.get("conditional_formatting", []):
        add_conditional_formatting(ws, rule_spec)

    for dv_spec in sheet_spec.get("data_validations", []):
        add_data_validation(ws, dv_spec)


def fill_template(template_path, output_path, data):
    if not os.path.isfile(template_path):
        raise ValueError(f"template 文件不存在: {template_path}")

    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.__class__.__name__ == "MergedCell":
                    continue
                if isinstance(cell.value, str) and "${" in cell.value:
                    new_value = cell.value
                    for key, val in data.items():
                        new_value = new_value.replace("${" + key + "}", str(val))
                    cell.value = new_value

    wb.save(output_path)

    if not os.path.isfile(output_path):
        raise RuntimeError(f"保存后文件不存在: {output_path}")
    try:
        load_workbook(output_path)
    except Exception as exc:
        raise RuntimeError(f"保存后 openpyxl 无法读取生成的文件: {exc}")

    return output_path


def generate(spec):
    if not isinstance(spec, dict):
        raise ValueError("invalid spec: 根节点必须是 JSON object")

    output_path = spec.get("output")
    if not output_path:
        raise ValueError("missing required field: output")

    template_path = spec.get("template")
    if template_path:
        data = spec.get("data")
        if not isinstance(data, dict):
            raise ValueError("missing required field: data (must be object) when template is set")
        return fill_template(template_path, output_path, data)

    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("missing required field: sheets (must be non-empty list)")

    for sheet_spec in sheets:
        if not isinstance(sheet_spec, dict):
            raise ValueError("invalid sheet spec: 每个 sheet 必须是 JSON object")
        rows = sheet_spec.get("rows", [])
        if not isinstance(rows, list) or any(not isinstance(row, list) for row in rows):
            raise ValueError("invalid sheet spec: rows 必须是二维数组")

    wb = Workbook()
    wb.remove(wb.active)

    used_names = [sheet_spec.get("name") for sheet_spec in sheets if sheet_spec.get("name")]
    for sheet_spec in sheets:
        name = sheet_spec.get("name")
        if not name:
            name = "Sheet"
            suffix = 2
            while name in used_names:
                name = f"Sheet{suffix}"
                suffix += 1
            used_names.append(name)
        ws = wb.create_sheet(title=name)
        write_sheet(ws, sheet_spec)

    wb.save(output_path)

    if not os.path.isfile(output_path):
        raise RuntimeError(f"保存后文件不存在: {output_path}")
    try:
        load_workbook(output_path)
    except Exception as exc:
        raise RuntimeError(f"保存后 openpyxl 无法读取生成的文件: {exc}")

    return output_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", help="JSON 输入文件路径；不传则从 stdin 读取")
    args = parser.parse_args()

    try:
        if args.input:
            with open(args.input, "r", encoding="utf-8") as f:
                raw = f.read()
        else:
            raw = sys.stdin.read()
    except OSError as exc:
        error_out("io_error", f"无法读取输入: {exc}")

    try:
        spec = json.loads(raw)
    except json.JSONDecodeError as exc:
        error_out("invalid_json", f"JSON 解析失败: {exc}")

    try:
        output_path = generate(spec)
    except ValueError as exc:
        error_out("invalid_spec", str(exc))
    except Exception as exc:
        error_out("generation_error", str(exc), traceback.format_exc())

    print(json.dumps({"ok": True, "output": output_path}, ensure_ascii=False))


if __name__ == "__main__":
    main()
